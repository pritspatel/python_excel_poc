from openpyxl import load_workbook

# Load the excel file
wb = load_workbook(filename = 'data/Financial Sample.xlsx')

# Find worksheets
ws = wb.active

# Display total cols
print(ws.max_column)
#Display total row
print(ws.max_row)

# # Display all col names
# for i in range(1, ws.max_column + 1):
#    my_cell_obj = ws.cell(row = 1, column = i)
#    print(my_cell_obj.value)

# Dsiplay all rows
# for i in range(1, ws.max_row + 1):
#    cell_obj = ws.cell(row = i, column = 1)
#    print(cell_obj.value)

# Display particular row value
for i in range(1, ws.max_column + 1):
   cell_obj = ws.cell(row = 2, column = i)
   print(cell_obj.value, end = " ")